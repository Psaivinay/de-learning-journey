    
import openpyxl
import csv
import logging

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Sample data
pipeline_data = [
    {"pipeline": "NYC Taxi Bronze", "status": "completed", "records": 10000},
    {"pipeline": "NYC Taxi Silver", "status": "failed",    "records": 0},
    {"pipeline": "NYC Taxi Gold",   "status": "completed", "records": 100},
]

# ── STEP 1: Write to CSV ──
with open('pipelines.csv', 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=["pipeline", "status", "records"])
    writer.writeheader()
    writer.writerows(pipeline_data)

logging.info("CSV written successfully")

# ── STEP 2: Read from CSV ──
with open('pipelines.csv', 'r') as f:
    reader = csv.DictReader(f)
    for row in reader:
        print(row)

logging.info("CSV read successfully")

# ── STEP 3: Filter failed runs ──
with open('pipelines.csv', 'r') as f:
    reader = csv.DictReader(f)
    failed = [row for row in reader if row['status'] == 'failed']

print(f"\nFailed runs: {len(failed)}")
for f in failed:
    print(f"  → {f['pipeline']}")

# ── STEP 4: Write to Excel ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pipeline Runs"

# Write header row
headers = ["pipeline", "status", "records"]
for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header.upper())

# Write data rows
for row_num, record in enumerate(pipeline_data, start=2):
    ws.cell(row=row_num, column=1, value=record["pipeline"])
    ws.cell(row=row_num, column=2, value=record["status"])
    ws.cell(row=row_num, column=3, value=record["records"])

wb.save("pipelines.xlsx")
logging.info("Excel file saved successfully")

# ── STEP 5: Read from Excel ──
wb2 = openpyxl.load_workbook("pipelines.xlsx")
ws2 = wb2.active

for row in ws2.iter_rows(min_row=2, values_only=True):
    print(row)

logging.info("Excel read successfully")